import pandas as pd
from io import StringIO
from openpyxl import Workbook
import re

# Read your CSV file into a pandas DataFrame
# Replace 'path/to/your/input/file.csv' with the actual path to your CSV file
input_csv_file = r'E:\Users\noero\Desktop\order.csv'
data = pd.read_csv(input_csv_file, delimiter=';')

# Create an empty list to store dictionaries for each sweatshirt
sweatshirts_data = []

# Iterate through each row in the DataFrame
for index, row in data.iterrows():
    # Create a dictionary to store information for the current sweatshirt
    sweatshirt_info = {
        'name': row['name'],
        'quantity': row['quantity']
    }

    # Extract information from the 'options' column
    options = row['options'].split("\n")

    # Extract information from each line in 'options'
    for line in options:
        key, value = line.split(":", 1)
        sweatshirt_info[key.strip()] = value.strip()

    # Append the dictionary to the list
    sweatshirts_data.append(sweatshirt_info)

# Create a DataFrame from the list of dictionaries
df = pd.DataFrame(sweatshirts_data)





# Assuming df is your DataFrame with the structure from the provided data
# If you are using the previous code snippet, make sure you have the df DataFrame.

# Function to clean and format column headers
def clean_header(header_str):
    cleaned_str = re.sub(r'[^a-zA-Z0-9\s]', '', header_str)  # Remove special characters
    cleaned_str = cleaned_str.lower().replace(' ', '_')  # Convert to lowercase and replace spaces with underscores
    return cleaned_str

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Create a header row
header = ['Logo', 'Couleur', 'Total']

# Iterate through each row in the DataFrame
for index, row in df.iterrows():
    logo = row['Logo']
    couleur = row['Couleur']
    taille = row['Taille']
    type_value = row['Type']
    couleur_broderie_flocage = row['Couleur Broderie/Flocage']
    quantity = row['quantity']

    # Remove extra details from 'Type' value and clean headers
    type_value_cleaned = re.sub(r'[^a-zA-Z0-9\s]', '', type_value.split('(')[0].strip()).lower().replace(' ', '_')
    couleur_broderie_flocage_cleaned = couleur_broderie_flocage.split('(')[0].strip()

    # Construct the expected column header
    column_header = f'{taille.lower()}_{type_value_cleaned}_{couleur_broderie_flocage_cleaned}'

    # Check if the header is already in the Excel sheet
    if column_header not in header:
        header.append(column_header)
        ws.cell(row=1, column=len(header), value=column_header)

    # Write information to the Excel sheet
    ws.cell(row=index + 2, column=header.index('Logo') + 1, value=logo)
    ws.cell(row=index + 2, column=header.index('Couleur') + 1, value=couleur)
    ws.cell(row=index + 2, column=header.index(column_header) + 1, value=quantity)

# Save the Excel workbook
excel_file = 'output.xlsx'
wb.save(excel_file)